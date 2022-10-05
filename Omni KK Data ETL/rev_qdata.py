#!/usr/bin/env Python3

import PD_ETL as dQ
import openpyxl
import sys
import os
import shutil

def main(args):     
    # args is intended to be strings of
    # 1. file name to be processed
    # 2. model code "RC153" "RC265-04B" "RC305"
    # 3~. array of dates of updated data in format wwww.d wwww.d ...
    ETL_files = dQ.file_list();
    arch_f = "archive"
    cnxn = dQ.SQL_init();
    cursor = cnxn.cursor()
    for file in ETL_files:
        if file != args[1]:
            continue
        xlsx = openpyxl.load_workbook(file)
        for upDate in args[3:]:
            print(upDate[0:4]);
            dQ.logger(3, "started processing {} to update data on {}".format(args[1], upDate))
            cursor.execute("SELECT ID FROM datetable WHERE \
                weeknum = " + upDate[0:4] + " AND dayofweek = " + upDate[-1])
            SQLdata = cursor.fetchone()
            dateID = SQLdata[0]
            modelID = dQ.model_ID(args[2])
            cursor.execute("DELETE FROM dailyPD WHERE dateID = {} AND modelID = {};".format(dateID, modelID))
            cursor.execute("DELETE FROM dailydefects WHERE dateID = {} AND modelID = {};".format(dateID, modelID))
            for sheet in xlsx.sheetnames:
                xlsx.active = xlsx.sheetnames.index(sheet)
                asheet = xlsx.active
                icolumn = dQ.locate_data(asheet)
                if icolumn == 'NDAT':
                    continue
                datecode = asheet.cell(row = 1, column = icolumn).value[2:]
                while datecode != upDate:
                    icolumn -= 1;
                    if asheet.cell(row = 1, column = icolumn).value[-1].isalpha():
                        icolumn -= 1;
                    datecode = asheet.cell(row = 1, column = icolumn).value[2:]
                    if icolumn < 5 or int(datecode[:4]) < int(upDate[:4]):
                        break;
                if icolumn < 5 or int(datecode[:4]) < int(upDate[:4]):
                    continue
                icolumn_letter = openpyxl.utils.cell.get_column_letter(icolumn)
                dcolumn = asheet[icolumn_letter]
                PDdata = {}
                PDdata['PDinput'] = dcolumn[2].value
                PDdata['PDoutput'] = dcolumn[3].value
                PDdata['modelID'] = dQ.model_ID(dcolumn[1].value)
                if PDdata['modelID'] != modelID:
                    continue
                PDdata['pdtypeID'] = dQ.production_type(sheet, PDdata['modelID'])
                if (type(PDdata['modelID']) is str) or (type(PDdata['pdtypeID']) is str):
                    dQ.logger(1, "model or PD type unrecognized.")
                # read defect data
                defects = dQ.defect_reader(asheet, dcolumn, PDdata['pdtypeID'], cursor)
                dQ.insert_data(cursor, dateID, PDdata, defects)
            cnxn.commit()
            dQ.logger(0, "{} data on {} ({}) is successfully updated.".format(file, upDate, args[2]))
        xlsx.close()
        shutil.move(file, os.path.join(arch_f, file))
    cnxn.close()
    dQ.trigger_refresh()
    dQ.logger(3, "finished processing")

if __name__ == '__main__':
    main(sys.argv)