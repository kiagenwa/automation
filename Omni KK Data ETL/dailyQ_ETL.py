#!/usr/bin/env Python3

import os
import shutil
import openpyxl
import PD_ETL as dQ

ETL_files = dQ.file_list();
arch_f = "archive"
cnxn = dQ.SQL_init();
cursor = cnxn.cursor()
dQ.logger(3, "started processing")

# process each file
for file in ETL_files:
    if file.split('.')[-1] != "xlsx":
        continue
    xlsx = openpyxl.load_workbook(file)
    if not xlsx.sheetnames:
        continue
    dateID = dQ.get_latest_dateID(xlsx, cursor)
    for sheet in xlsx.sheetnames:
        xlsx.active = xlsx.sheetnames.index(sheet)
        asheet = xlsx.active
        icolumn = dQ.locate_data(asheet)
        if icolumn == 'NDAT':
            continue
        if (dateID > dQ.get_dateID(asheet, icolumn, cursor)):
            continue
        icolumn_letter = openpyxl.utils.cell.get_column_letter(icolumn)
        dcolumn = asheet[icolumn_letter]
        PDdata = {}
        PDdata['PDinput'] = dcolumn[2].value
        PDdata['PDoutput'] = dcolumn[3].value
        PDdata['modelID'] = dQ.model_ID(dcolumn[1].value)
        PDdata['pdtypeID'] = dQ.production_type(sheet, PDdata['modelID'])
        if (type(PDdata['modelID']) is str) or (type(PDdata['pdtypeID']) is str):
            dQ.logger(1, "model or PD type unrecognized.")
        cursor.execute("SELECT dateID, modelID, pdtypeID FROM dailyPD \
            WHERE dateID = {} AND modelID = {} AND pdtypeID = \
                {}".format(dateID, PDdata['modelID'], PDdata['pdtypeID']))
        SQLdata = cursor.fetchall()
        if SQLdata:
            dQ.logger(1, "{} in {} is already recorded in warehouse.".format(sheet, file))
        # read defect data
        defects = dQ.defect_reader(asheet, dcolumn, PDdata['pdtypeID'], cursor)
        dQ.insert_data(cursor, dateID, PDdata, defects)
    cnxn.commit()
    dQ.logger(0, "{} is successfully loaded".format(file))
    xlsx.close()
    shutil.move(file, os.path.join(arch_f, file))

cnxn.close()
dQ.trigger_refresh()
dQ.logger(3, "finished processing")