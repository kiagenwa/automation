#!/usr/bin/env Python3

import os
import sys
import shutil
import re
import subprocess
import pyodbc
import openpyxl
from datetime import datetime

def logger(flag, message):
    """appends message to a log file and exit the program"""
    LogFile = "C:\\Users\\KH1000\\OneDrive - OMNI Remotes\\technical stuff\\ETL\\log\\ETL_log.txt"
    switch={
        0:"SUCCESS ",
        1:"FAILED ",
        2:"WARN ",
        3: "INFO "
    }
    with open(LogFile, "a") as log_file:
        log_file.write(str(datetime.now()) + "\t" + switch[flag] + message + "\n")
    if flag == 1:
        sys.exit(message)
    pass

def model_ID(model_abbr):
    """returns model ID with model abbreviation"""
    switch={
        "RC153":5,
        "RC265-04B":1,
        "RC305":4
    }
    return switch.get(model_abbr, "AA")

def production_type(sheetname, modelID):
    """returns pdtype ID with sheet name"""
    if re.search(r"main",sheetname.lower()) or (re.search(r'smt|smd',sheetname.lower()) and modelID == 1):
        return 4
    matched = re.search(r'smt|smd|ws|assembly|wave',sheetname.lower())
    switch={
        "assembly":3,
        "smt":1,
        "smd":1,
        "ws":2,
        "wave":2
    }
    return switch.get(matched[0], "BB")

def locate_data(sheet_obj):
    """returns the column number of the latest data"""
    lcolumn = sheet_obj.max_column - 1
    while not sheet_obj.cell(row = 3, column = lcolumn).value:
        lcolumn -= 1
        #skip a column when the max column is not the total column
        test = sheet_obj.cell(row = 1, column = lcolumn).value
        if not test:
            lcolumn -= 1
        elif not test.strip()[-1].isdigit():
            lcolumn -= 1
    return lcolumn

def get_dateID(sheet_obj, icolumn, cursor):
    """returns dateID for the sheet's latest data"""
    datecode = sheet_obj.cell(row = 1, column = icolumn).value
    wweek = datecode[2:6]
    wday = datecode[-1]
    try:
        int(wday)
    except:
        return 1
    cursor.execute("SELECT ID FROM datetable WHERE \
        weeknum = " + wweek + " AND dayofweek = " + wday)
    SQLdata = cursor.fetchone()
    if not SQLdata:
        logger(1, "No date ID available for this report. \
            Please update datetable.")
    dateID = SQLdata[0]
    return dateID

def get_latest_dateID(xlsx, cursor):
    """returns latest dateID for the report file"""
    dateIDs = []
    for sheet in xlsx.sheetnames:
        xlsx.active = xlsx.sheetnames.index(sheet)
        asheet = xlsx.active
        icolumn = locate_data(asheet)
        dateIDs.append(get_dateID(asheet, icolumn, cursor))
    dateIDs.sort()
    return dateIDs[-1]

def defect_reader(asheet, dcolumn, pdtypeID):
    """returns a dictionary of defect ID:qty"""
    # 1=solder SMD, 2=WS, 3=assembly, 4=glue SMD
    stp_switch={
        1:29,
        2:27,
        3:17,
        4:25
    }
    idx_switch={
        1:-48,
        2:-15,
        3:13,
        4:-32
    }
    st_switch={
        1:11,
        2:11,
        3:14,
        4:11
    }
    defects = {}
    for def_idx in range(st_switch[pdtypeID], stp_switch[pdtypeID]):
        qty = dcolumn[def_idx-1].value
        try:
            defects[(def_idx - idx_switch[pdtypeID])] = int(qty)
        except:
            continue
    if pdtypeID == 3:
        st = 22
        stp = 42
        idx = 17
        if asheet.cell(row = 20, column = 1).value:
            st -= 1
            stp -= 1
            idx -= 1
        for def_idx in range(st, stp):
            qty = dcolumn[def_idx-1].value
            if qty:
                defects[(def_idx - idx)] = qty
    return defects

def insert_data(cursor, dateID, PDdata, defects):
    """collects and inserts data into SQL server"""
    cursor.execute("INSERT INTO dailyPD (dateID, PDinput, PDoutput, \
        modelID, pdtypeID) VALUES ({},{},{},{},{});".format(dateID, 
        PDdata['PDinput'], PDdata['PDoutput'], PDdata['modelID'], PDdata['pdtypeID']))
    if defects:
        insert_stm = "INSERT INTO dailydefects (dateID, defectID, \
            qty, modelID) VALUES "
        for defectID in defects.keys():
            insert_stm += "({},{},{},{}), ".format(dateID,
            defectID, defects[defectID], PDdata['modelID'])
        cursor.execute(insert_stm[:-2])
    pass

def trigger_refresh():
    """create a file in trigger folder to have Power BI data refreshed"""
    tref = str(datetime.now())
    os.chdir("trigger")
    #subprocess.call("del *.* /q",shell=True)
    subprocess.call("type nul >> {}-{}_trigger.m".format(tref[:10], tref[-6:]),shell=True)
    os.chdir("..")
    pass

ETLpath = "C:\\Users\\KH1000\\OneDrive - OMNI Remotes\\technical stuff\\ETL"
arch_f = "archive"
ETL_files = [f for f in os.listdir(ETLpath) if os.path.isfile(os.path.join(ETLpath, f))]
os.chdir(ETLpath)
if not ETL_files:
    logger(1, "No files to process")

# initialize SQL connection
server = "kh1000\jr_omni"
database = "Daily_Quality"
username = "blank"          #blanked for purpose of publication
password = "blank"
cnxn = pyodbc.connect("DRIVER={ODBC Driver 17 for SQL Server}; \
    SERVER="+server+";DATABASE="+database+";UID="+username+";PWD="+password)
cursor = cnxn.cursor()

logger(3, "started processing")

# process each file
for file in ETL_files:
    if file.split('.')[-1] != "xlsx":
        continue
    xlsx = openpyxl.load_workbook(file)
    if not xlsx.sheetnames:
        continue
    dateID = get_latest_dateID(xlsx, cursor)
    for sheet in xlsx.sheetnames:
        xlsx.active = xlsx.sheetnames.index(sheet)
        asheet = xlsx.active
        icolumn = locate_data(asheet)
        if (icolumn < 5) or (dateID > get_dateID(asheet, icolumn, cursor)):
            continue
        icolumn_letter = openpyxl.utils.cell.get_column_letter(icolumn)
        dcolumn = asheet[icolumn_letter]
        PDdata = {}
        PDdata['PDinput'] = dcolumn[2].value
        PDdata['PDoutput'] = dcolumn[3].value
        PDdata['modelID'] = model_ID(dcolumn[1].value)
        PDdata['pdtypeID'] = production_type(sheet, PDdata['modelID'])
        if (type(PDdata['modelID']) is str) or (type(PDdata['pdtypeID']) is str):
            logger(1, "model or PD type unrecognized.")
        cursor.execute("SELECT dateID, modelID, pdtypeID FROM dailyPD \
            WHERE dateID = {} AND modelID = {} AND pdtypeID = \
                {}".format(dateID, PDdata['modelID'], PDdata['pdtypeID']))
        SQLdata = cursor.fetchall()
        if SQLdata:
            logger(1, "{} in {} is already recorded in warehouse.".format(sheet, file))
        # read defect data
        defects = defect_reader(asheet, dcolumn, PDdata['pdtypeID'])
        insert_data(cursor, dateID, PDdata, defects)
    cnxn.commit()
    logger(0, "{} is successfully loaded".format(file))
    xlsx.close()
    shutil.move(os.path.join(ETLpath, file), os.path.join(ETLpath, arch_f, file))

cnxn.close()
trigger_refresh()
logger(3, "finished processing")